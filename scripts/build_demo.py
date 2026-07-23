# -*- coding: utf-8 -*-
"""
D365 Finance & Supply Chain -- Payment Prediction demo builder.

One command does everything:

    python scripts/build_demo.py

  * If data/PaymentPrediction-Demo.xlsx does not exist, it is created from the
    seed pack (default: the South Bend International Airport pack in seeds.py).
  * Reads the `Config` and `Input_Customers` sheets -- these are YOURS to edit.
  * Regenerates every output sheet (Transactions, CustomerPredictions,
    AgedBalances, TopFactors, Customers) deterministically from those inputs.
  * Rebuilds dashboard/payment-prediction.html with the data baked in.

To retarget the demo at a different customer / industry: edit `Input_Customers`
in the workbook (or drop in a whole new workbook with the same sheet layout),
then re-run this script. Nothing else changes.

Flags
    --seed <pack>   Seed pack to use when creating a new workbook (default: airport)
    --force-seed    Overwrite Config + Input_Customers from the seed pack
    --workbook <p>  Path to the workbook       (default: data/PaymentPrediction-Demo.xlsx)
    --out <p>       Path to the dashboard HTML (default: dashboard/payment-prediction.html)
    --no-html       Regenerate the workbook only
"""

from __future__ import annotations

import argparse
import json
import math
import random
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from seeds import BILLING_LINES, SEED_PACKS  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = ROOT / "data" / "PaymentPrediction-Demo.xlsx"
DEFAULT_HTML = ROOT / "dashboard" / "payment-prediction.html"
TEMPLATE = ROOT / "dashboard" / "template.html"

INPUT_SHEETS = ("Config", "Input_Customers")
OUTPUT_SHEETS = ("Customers", "Transactions", "CustomerPredictions",
                 "AgedBalances", "TopFactors", "README")

TERMS_DAYS = {"Net 15": 15, "Net 30": 30, "Net 45": 45, "Net 60": 60, "COD": 0}

# Days-past-due pools per aging profile. Index i is used for the i-th open item.
# Ordered so that even a customer with only 3-4 open items lands across several
# aged-balance buckets, rather than piling into one.
AGING_POOLS = {
    "current":  [-28, -14, -21, -7, -2],
    "mild":     [-9, 6, -18, 21, 12, 33],
    "moderate": [16, 38, 67, 27, 49, 83],
    "severe":   [45, 78, 119, 196, 152, 88, 61, 232],
}

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10, name="Segoe UI")
BODY_FONT = Font(size=10, name="Segoe UI")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = '#,##0.00'
PCT = '0%'
DATEFMT = 'm/d/yyyy'

# Colour themes offered by D365 F&O under User options > Visual > Color theme,
# in the order they appear there, plus the brick red used by the current RSMUS
# environment. The dashboard's own User options page renders the same set.
THEMES = ["blue", "berry", "orchid", "orange", "green", "darkblue", "darkpurple",
          "azure", "olive", "pink", "teal", "steel", "darkred", "maroon",
          "highcontrast", "red"]

CONFIG_NOTES = {
    "Demo name": "Your own label for this scenario",
    "Legal entity": "Rendered in the F&O header chip, e.g. SBN",
    "Organization": "Legal entity long name shown next to the chip",
    "User name": "Demo persona shown on the avatar tooltip",
    "User initials": "Avatar initials. Leave blank to derive from User name",
    "User email": "Display only - the dashboard has nothing to sign in to",
    "Currency": "Applied to every transaction",
    "As of date": "The 'today' predictions are calculated against (YYYY-MM-DD)",
    "Aging period definition": "Label in the Aged balances FactBox",
    "Red dot threshold (%)": "On-time probability below this shows the red dot",
    "Theme": "One of: " + ", ".join(THEMES),
    "Size": "comfortable or compact - element density, as in D365 User options",
    "Random seed": "Change this to reshuffle amounts/probabilities",
}

# Defaults back-filled into workbooks created before a setting existed.
CONFIG_DEFAULTS = {
    "User name": "Demo User",
    "User initials": "",
    "User email": "demo.user@contoso.com",
    "Size": "comfortable",
}


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------
def sigmoid(x: float) -> float:
    if x < -60:
        return 0.0
    if x > 60:
        return 1.0
    return 1.0 / (1.0 + math.exp(-x))


def clamp(v, lo, hi):
    return max(lo, min(hi, v))


def as_date(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v).strip()[:10], "%Y-%m-%d").date()


# ---------------------------------------------------------------------------
# core generation
# ---------------------------------------------------------------------------
def split_probabilities(risk: float, dpd: int, amount: float, rng: random.Random):
    """Return (on_time, late, very_late) as integer percentages summing to 100."""
    z = (3.4 * risk
         - 1.70
         - 0.045 * max(dpd, 0)
         + (0.60 if dpd < 0 else 0.0)
         - 0.28 * math.log10(max(amount, 1.0) / 10_000.0)
         + rng.gauss(0, 0.32))
    on_time = clamp(sigmoid(z), 0.01, 0.98)

    vl_share = clamp(0.02 + 0.006 * max(dpd, 0) - 0.20 * risk, 0.0, 0.45)
    very_late = (1.0 - on_time) * vl_share

    ot = int(round(on_time * 100))
    vl = int(round(very_late * 100))
    ot = clamp(ot, 1, 98)
    vl = clamp(vl, 0, 99 - ot)
    lt = 100 - ot - vl
    return ot, lt, vl


def generate(customers: list[dict], config: dict):
    """Build transactions + per-customer rollups from the customer input rows."""
    rng = random.Random(int(config.get("Random seed", 20260723)))
    as_of = as_date(config.get("As of date", date.today().isoformat()))
    currency = str(config.get("Currency", "USD"))

    counters = {"Customer": 30000001, "Collection letter": 50000001,
                "Project": 180000101, "Sales order": 10000701, "Interest note": 70000001}
    inv_counters = {"Customer": 1002, "Collection letter": 20, "Project": 101,
                    "Sales order": 715, "Interest note": 41}

    transactions: list[dict] = []
    cust_rows: list[dict] = []

    for cust in customers:
        group = cust["group"]
        lines = BILLING_LINES.get(group)
        if not lines:
            lines = [("Miscellaneous billing", "Customer", 1_000, 25_000)]
        pool = AGING_POOLS.get(str(cust.get("aging", "mild")).lower(), AGING_POOLS["mild"])
        terms = str(cust.get("terms", "Net 30"))
        terms_days = TERMS_DAYS.get(terms, 30)
        risk = float(cust["risk"])
        opens = int(cust.get("opens", 3))

        my_tx: list[dict] = []
        for i in range(opens):
            desc, ttype, lo, hi = lines[i % len(lines)]
            dpd = pool[i % len(pool)] + rng.randint(-3, 3)
            due = as_of - timedelta(days=dpd)
            inv_date = due - timedelta(days=terms_days)
            amount = round(rng.uniform(lo, hi), 2)
            balance = amount
            if rng.random() < 0.12:                      # partially settled
                balance = round(amount * rng.uniform(0.35, 0.85), 2)

            ot, lt, vl = split_probabilities(risk, dpd, amount, rng)

            v = counters[ttype]
            counters[ttype] += rng.randint(3, 9)
            n = inv_counters[ttype]
            inv_counters[ttype] += rng.randint(2, 7)

            if ttype == "Customer":
                voucher, invoice = f"FTV-{v}", f"FTI-{n:08d}"
            elif ttype == "Collection letter":
                voucher, invoice = f"CLV-{v}", f"{n:06d}"
            elif ttype == "Project":
                voucher, invoice = f"{v}", f"{n:06d}"
            elif ttype == "Sales order":
                voucher, invoice = f"INV-{v}", f"CIV-{n:06d}"
            else:
                voucher, invoice = f"INT-{v}", f"{n:06d}"

            my_tx.append(dict(
                account=cust["account"], name=cust["name"], group=group,
                on_time=ot, late=lt, very_late=vl,
                voucher=voucher, invoice=invoice, ttype=ttype, description=desc,
                date=inv_date, due=due, dpd=dpd,
                amount=amount, balance=balance, currency=currency, terms=terms,
            ))

        # Delinquent accounts pick up a collection letter.
        if risk < 0.50 and str(cust.get("aging", "")).lower() == "severe":
            worst = max(my_tx, key=lambda t: t["dpd"])
            v = counters["Collection letter"]; counters["Collection letter"] += rng.randint(3, 9)
            n = inv_counters["Collection letter"]; inv_counters["Collection letter"] += rng.randint(2, 7)
            amount = round(worst["balance"] * rng.uniform(0.02, 0.05) + rng.uniform(15, 60), 2)
            dpd = max(worst["dpd"] - rng.randint(20, 45), 1)
            due = as_of - timedelta(days=dpd)
            ot, lt, vl = split_probabilities(risk * 0.7, dpd, amount, rng)
            my_tx.append(dict(
                account=cust["account"], name=cust["name"], group=group,
                on_time=ot, late=lt, very_late=vl,
                voucher=f"CLV-{v}", invoice=f"{n:06d}", ttype="Collection letter",
                description="Collection letter fee - 2nd notice",
                date=due - timedelta(days=terms_days), due=due, dpd=dpd,
                amount=amount, balance=amount, currency=currency, terms=terms,
            ))

        my_tx.sort(key=lambda t: t["due"])
        transactions.extend(my_tx)

        # ---- per-customer rollups -----------------------------------------
        open_amount = sum(t["balance"] for t in my_tx)
        ot_amt = sum(t["balance"] * t["on_time"] / 100 for t in my_tx)
        lt_amt = sum(t["balance"] * t["late"] / 100 for t in my_tx)
        vl_amt = sum(t["balance"] * t["very_late"] / 100 for t in my_tx)

        ot_pct = int(round(ot_amt / open_amount * 100)) if open_amount else 0
        vl_pct = int(round(vl_amt / open_amount * 100)) if open_amount else 0
        ot_pct = clamp(ot_pct, 0, 100)
        vl_pct = clamp(vl_pct, 0, 100 - ot_pct)
        lt_pct = 100 - ot_pct - vl_pct

        # Collection letters are their own document type in D365 -- they are
        # counted as collection cases, never as invoices.
        invoices = [t for t in my_tx if t["ttype"] != "Collection letter"]
        late_tx = [t for t in invoices if t["dpd"] > 0]
        collection_cases = sum(1 for t in my_tx if t["ttype"] == "Collection letter")

        avg_dpd = sum(max(t["dpd"], 0) for t in my_tx) / len(my_tx)
        avg_days_to_pay = int(round(terms_days + avg_dpd * 0.75 + (1 - risk) * 22))

        open_activities = 0 if risk > 0.8 else rng.randint(0, 1 if risk > 0.55 else 4)

        last_pay_days = int(round(18 + (1 - risk) * 140 + rng.randint(-6, 6)))
        last_pay_date = as_of - timedelta(days=last_pay_days)
        last_pay_amount = round(open_amount * rng.uniform(0.15, 0.65), 2)

        avg_invoice_balance = round(open_amount / max(len(invoices), 1), 2)
        avg_amount_late = round(sum(t["balance"] for t in late_tx) / len(late_tx), 2) if late_tx else 0.0

        credit_limit = max(5_000, int(round(open_amount * rng.uniform(1.5, 3.0) / 5_000)) * 5_000)

        # aged balances
        buckets = {"Current": 0.0, "30 days": 0.0, "60 days": 0.0,
                   "90 days": 0.0, "180 and over": 0.0}
        for t in my_tx:
            d = t["dpd"]
            if d <= 0:
                buckets["Current"] += t["balance"]
            elif d <= 30:
                buckets["30 days"] += t["balance"]
            elif d <= 60:
                buckets["60 days"] += t["balance"]
            elif d <= 90:
                buckets["90 days"] += t["balance"]
            else:
                buckets["180 and over"] += t["balance"]

        # top factors -- mirrors the "Top factors" list in the D365 FactBox
        avg_amount = open_amount / max(len(my_tx), 1)
        scored = [
            ("Average days to pay", avg_dpd / 25.0 + (1 - risk) * 1.3),
            ("Customer group", 0.95),
            ("Amount", math.log10(max(avg_amount, 1) / 10_000.0) + 0.75),
            ("Payment terms", terms_days / 55.0),
            ("Open collection cases", 1.60 if collection_cases else 0.0),
            ("Number of open invoices", len(my_tx) / 5.5),
            ("Credit limit utilization", clamp(open_amount / credit_limit, 0, 1.4)),
        ]
        scored.sort(key=lambda x: -x[1])
        factors = [f[0] for f in scored[:3]]

        cust_rows.append(dict(
            account=cust["account"], name=cust["name"], group=group, terms=terms,
            currency=currency, risk=risk, aging=cust.get("aging", "mild"),
            city=cust.get("city", ""), address=cust.get("address", ""),
            balance=round(open_amount, 2),
            on_time_pct=ot_pct, late_pct=lt_pct, very_late_pct=vl_pct,
            on_time_amt=round(ot_amt, 2), late_amt=round(lt_amt, 2), very_late_amt=round(vl_amt, 2),
            open_invoices=len(invoices), late_invoices=len(late_tx),
            collection_cases=collection_cases, open_activities=open_activities,
            last_payment_amount=last_pay_amount, last_payment_date=last_pay_date,
            avg_invoice_balance=avg_invoice_balance, avg_amount_late=avg_amount_late,
            avg_days_to_pay=avg_days_to_pay, credit_limit=credit_limit,
            credit_used_pct=int(round(clamp(open_amount / credit_limit, 0, 1) * 100)),
            aged=buckets, factors=factors,
        ))

    cust_rows.sort(key=lambda c: c["account"])
    transactions.sort(key=lambda t: (t["account"], t["due"]))
    return cust_rows, transactions


# ---------------------------------------------------------------------------
# workbook I/O
# ---------------------------------------------------------------------------
def style_sheet(ws, widths, freeze="A2", n_cols=None):
    n_cols = n_cols or len(widths)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1][:n_cols]:
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = freeze
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{ws.max_row}"


def write_input_sheets(wb, pack):
    cfg = wb.create_sheet("Config", 0)
    cfg.append(["Setting", "Value", "Notes"])
    for k, v in pack["config"].items():
        cfg.append([k, v, CONFIG_NOTES.get(k, "")])
    for row in cfg.iter_rows(min_row=2, max_row=cfg.max_row, max_col=3):
        for c in row:
            c.font, c.border = BODY_FONT, BORDER
        row[1].fill = INPUT_FILL
    style_sheet(cfg, [30, 34, 62])

    ci = wb.create_sheet("Input_Customers", 1)
    headers = ["Customer account", "Name", "Customer group", "Payment terms",
               "Risk (0-1)", "Aging profile", "Open transactions", "City", "Address"]
    ci.append(headers)
    for c in pack["customers"]:
        ci.append([c["account"], c["name"], c["group"], c["terms"], c["risk"],
                   c["aging"], c["opens"], c.get("city", ""), c.get("address", "")])
    for row in ci.iter_rows(min_row=2, max_row=ci.max_row, max_col=len(headers)):
        for c in row:
            c.font, c.border, c.fill = BODY_FONT, BORDER, INPUT_FILL
        row[4].number_format = '0.00'
        row[8].alignment = Alignment(wrap_text=True, vertical="top")
    style_sheet(ci, [18, 44, 24, 14, 10, 14, 12, 18, 34])


def seed_workbook(path: Path, pack_name: str):
    pack = SEED_PACKS[pack_name]
    wb = Workbook()
    wb.remove(wb.active)
    write_input_sheets(wb, pack)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return wb


def read_inputs(path: Path):
    """Return (workbook_to_write_into, config, customers).

    Two loads on purpose: `data_only` gives us computed values, but strips
    formulas on save. We write into a second, formula-preserving handle so a
    user can drive Input_Customers with formulas without losing them on rebuild.
    """
    vals = load_workbook(path, data_only=True)
    wb = load_workbook(path)

    for required in INPUT_SHEETS:
        if required not in vals.sheetnames:
            raise SystemExit(f"{path.name} has no '{required}' sheet. "
                             f"Run with --force-seed to rebuild the input sheets.")

    cfg = {}
    for r in vals["Config"].iter_rows(min_row=2, values_only=True):
        if r[0] is not None:
            cfg[str(r[0]).strip()] = r[1]

    # Back-fill settings added after this workbook was created, so an existing
    # demo picks up new options without needing --force-seed.
    added = []
    for key, default in CONFIG_DEFAULTS.items():
        if key not in cfg:
            cfg[key] = default
            ws = wb["Config"]
            ws.append([key, default, CONFIG_NOTES.get(key, "")])
            row = ws[ws.max_row]
            for c in row:
                c.font, c.border = BODY_FONT, BORDER
            row[1].fill = INPUT_FILL
            added.append(key)
    if added:
        print("Added new Config settings to the workbook: " + ", ".join(added))

    theme = str(cfg.get("Theme", "blue")).strip().lower()
    if theme not in THEMES:
        print(f"Note: Theme '{cfg.get('Theme')}' is not a known theme, falling back to 'blue'.\n"
              f"      Valid themes: {', '.join(THEMES)}")
        cfg["Theme"] = "blue"

    customers, problems = [], []
    for i, r in enumerate(vals["Input_Customers"].iter_rows(min_row=2, values_only=True), start=2):
        if r[0] is None or str(r[0]).strip() == "":
            continue
        try:
            risk = float(r[4])
            opens = int(r[6])
        except (TypeError, ValueError):
            problems.append(f"  row {i} ({r[0]}): 'Risk (0-1)' and 'Open transactions' "
                            f"must be numbers, got {r[4]!r} and {r[6]!r}")
            continue
        if not 0.0 <= risk <= 1.0:
            problems.append(f"  row {i} ({r[0]}): 'Risk (0-1)' must be between 0 and 1, got {risk}")
            continue
        if opens < 1:
            problems.append(f"  row {i} ({r[0]}): 'Open transactions' must be at least 1, got {opens}")
            continue
        aging = str(r[5] or "mild").strip().lower()
        if aging not in AGING_POOLS:
            problems.append(f"  row {i} ({r[0]}): 'Aging profile' must be one of "
                            f"{', '.join(AGING_POOLS)}, got {r[5]!r}")
            continue
        customers.append(dict(
            account=str(r[0]).strip(), name=str(r[1]).strip(), group=str(r[2] or "").strip(),
            terms=str(r[3] or "Net 30").strip(), risk=risk, aging=aging, opens=opens,
            city=str(r[7] or ""), address=str(r[8] or ""),
        ))

    if problems:
        raise SystemExit("Input_Customers has problems:\n" + "\n".join(problems) +
                         "\n\nFix those rows and re-run. (If a cell holds a formula, open "
                         "the workbook in Excel and save it once so the value is cached.)")

    unknown = sorted({c["group"] for c in customers} - set(BILLING_LINES))
    if unknown:
        print("Note: no billing lines defined for these customer groups, using a generic "
              "line for each:\n  " + "\n  ".join(unknown) +
              "\n  (add them to BILLING_LINES in scripts/seeds.py for realistic invoices)")

    return wb, cfg, customers


def write_outputs(wb, path: Path, cfg, cust_rows, transactions, red_dot):
    for name in OUTPUT_SHEETS:
        if name in wb.sheetnames:
            wb.remove(wb[name])

    # --- Customers -------------------------------------------------------
    ws = wb.create_sheet("Customers")
    ws.append(["Customer account", "Name", "Customer group", "Balance", "Currency",
               "On time", "Late", "Very late", "On time amount", "Late amount",
               "Very late amount", "Payment terms", "Credit limit", "Credit used %",
               "Average days to pay", "Risk score", "Aging profile", "City"])
    for c in cust_rows:
        ws.append([c["account"], c["name"], c["group"], c["balance"], c["currency"],
                   c["on_time_pct"] / 100, c["late_pct"] / 100, c["very_late_pct"] / 100,
                   c["on_time_amt"], c["late_amt"], c["very_late_amt"], c["terms"],
                   c["credit_limit"], c["credit_used_pct"] / 100, c["avg_days_to_pay"],
                   c["risk"], c["aging"], c["city"]])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=18):
        for cell in row:
            cell.font, cell.border = BODY_FONT, BORDER
        for i in (3, 8, 9, 10, 12):
            row[i].number_format = MONEY
        for i in (5, 6, 7, 13):
            row[i].number_format = PCT
    style_sheet(ws, [18, 44, 24, 14, 9, 9, 9, 10, 15, 14, 15, 14, 13, 12, 14, 11, 13, 18])

    # --- Transactions ----------------------------------------------------
    ws = wb.create_sheet("Transactions")
    ws.append(["Customer account", "Name", "Customer group", "On time probability",
               "Late probability", "Very late probability", "Risk flag", "Voucher",
               "Invoice", "Transaction type", "Description", "Date", "Due date",
               "Days past due", "Amount in transaction currency", "Balance", "Currency",
               "Payment terms"])
    for t in transactions:
        ws.append([t["account"], t["name"], t["group"], t["on_time"] / 100,
                   t["late"] / 100, t["very_late"] / 100,
                   "At risk" if t["on_time"] < red_dot else "",
                   t["voucher"], t["invoice"], t["ttype"], t["description"],
                   t["date"], t["due"], t["dpd"], t["amount"], t["balance"],
                   t["currency"], t["terms"]])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=18):
        for cell in row:
            cell.font, cell.border = BODY_FONT, BORDER
        for i in (3, 4, 5):
            row[i].number_format = PCT
        for i in (11, 12):
            row[i].number_format = DATEFMT
        for i in (14, 15):
            row[i].number_format = MONEY
    style_sheet(ws, [18, 40, 22, 13, 12, 13, 10, 16, 16, 17, 40, 12, 12, 12, 18, 14, 9, 13])

    # --- CustomerPredictions (the FactBox counters) ----------------------
    ws = wb.create_sheet("CustomerPredictions")
    ws.append(["Customer account", "Name", "Open invoices", "Late invoices",
               "Open collection cases", "Open activities", "Last payment amount",
               "Last payment date", "Balance", "Average invoice balance",
               "Average amount late", "Average days to pay", "Credit limit",
               "Credit used %", "On time", "Late", "Very late"])
    for c in cust_rows:
        ws.append([c["account"], c["name"], c["open_invoices"], c["late_invoices"],
                   c["collection_cases"], c["open_activities"], c["last_payment_amount"],
                   c["last_payment_date"], c["balance"], c["avg_invoice_balance"],
                   c["avg_amount_late"], c["avg_days_to_pay"], c["credit_limit"],
                   c["credit_used_pct"] / 100, c["on_time_pct"] / 100,
                   c["late_pct"] / 100, c["very_late_pct"] / 100])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=17):
        for cell in row:
            cell.font, cell.border = BODY_FONT, BORDER
        for i in (6, 8, 9, 10, 12):
            row[i].number_format = MONEY
        row[7].number_format = DATEFMT
        for i in (13, 14, 15, 16):
            row[i].number_format = PCT
    style_sheet(ws, [18, 40, 13, 12, 17, 13, 16, 15, 14, 18, 17, 16, 13, 12, 9, 9, 10])

    # --- AgedBalances ----------------------------------------------------
    ws = wb.create_sheet("AgedBalances")
    ws.append(["Customer account", "Name", "Aging period definition", "As of date",
               "Current", "30 days", "60 days", "90 days", "180 and over", "Balance"])
    aging_def = str(cfg.get("Aging period definition", "30_60_90_180"))
    as_of = as_date(cfg.get("As of date", date.today().isoformat()))
    for c in cust_rows:
        a = c["aged"]
        ws.append([c["account"], c["name"], aging_def, as_of,
                   round(a["Current"], 2), round(a["30 days"], 2), round(a["60 days"], 2),
                   round(a["90 days"], 2), round(a["180 and over"], 2), c["balance"]])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=10):
        for cell in row:
            cell.font, cell.border = BODY_FONT, BORDER
        row[3].number_format = DATEFMT
        for i in range(4, 10):
            row[i].number_format = MONEY
    style_sheet(ws, [18, 40, 22, 13, 14, 14, 14, 14, 16, 14])

    # --- TopFactors ------------------------------------------------------
    ws = wb.create_sheet("TopFactors")
    ws.append(["Customer account", "Name", "Rank", "Factor"])
    for c in cust_rows:
        for i, f in enumerate(c["factors"], start=1):
            ws.append([c["account"], c["name"], i, f])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4):
        for cell in row:
            cell.font, cell.border = BODY_FONT, BORDER
    style_sheet(ws, [18, 40, 8, 30])

    # --- README ----------------------------------------------------------
    ws = wb.create_sheet("README")
    lines = [
        ("How to use this workbook", True),
        ("", False),
        ("This workbook drives BOTH the HTML dashboard replica and the Power BI report.", False),
        ("", False),
        ("YOU EDIT THESE TWO SHEETS:", True),
        ("  Config             - demo name, legal entity, currency, 'as of' date, theme.", False),
        ("  Input_Customers    - one row per customer. This is the whole demo.", False),
        ("", False),
        ("EVERYTHING ELSE IS REGENERATED - do not hand-edit:", True),
        ("  Customers, Transactions, CustomerPredictions, AgedBalances, TopFactors", False),
        ("", False),
        ("To retarget the demo at a new prospect:", True),
        ("  1. Replace the rows in Input_Customers with their tenants / customers.", False),
        ("  2. Set Config > Demo name, Legal entity, Organization, As of date.", False),
        ("  3. Run:  python scripts/build_demo.py", False),
        ("  4. Open dashboard/payment-prediction.html, or refresh the Power BI report.", False),
        ("", False),
        ("Input_Customers column guide:", True),
        ("  Risk (0-1)         1.00 = always pays on time. 0.20 = chronic late payer.", False),
        ("                     This is the single biggest driver of on-time probability.", False),
        ("  Aging profile      current | mild | moderate | severe", False),
        ("                     Controls how past-due their open items are, which drives", False),
        ("                     the aged balance buckets and the red risk dots.", False),
        ("  Customer group     Must match a group in scripts/seeds.py BILLING_LINES,", False),
        ("                     otherwise a generic billing line is used. The group also", False),
        ("                     appears as a prediction 'top factor'.", False),
        ("  Open transactions  How many open documents to generate for this customer.", False),
        ("", False),
        ("All figures in this workbook are SYNTHETIC and generated for demonstration.", True),
    ]
    for text, bold in lines:
        ws.append([text])
        ws.cell(ws.max_row, 1).font = Font(name="Segoe UI", size=10, bold=bold)
    ws.column_dimensions["A"].width = 100
    ws.sheet_view.showGridLines = False

    order = ["Config", "Input_Customers", "README", "Customers", "Transactions",
             "CustomerPredictions", "AgedBalances", "TopFactors"]
    wb._sheets = [wb[n] for n in order if n in wb.sheetnames]
    wb.save(path)


# ---------------------------------------------------------------------------
# dashboard build
# ---------------------------------------------------------------------------
def build_payload(cfg, cust_rows, transactions, red_dot):
    def d(x):
        return x.isoformat() if isinstance(x, (date, datetime)) else x

    return {
        "config": {
            "demoName": str(cfg.get("Demo name", "")),
            "legalEntity": str(cfg.get("Legal entity", "USMF")),
            "organization": str(cfg.get("Organization", "")),
            "userName": str(cfg.get("User name", "Demo User")),
            "userInitials": str(cfg.get("User initials") or ""),
            "userEmail": str(cfg.get("User email") or ""),
            "currency": str(cfg.get("Currency", "USD")),
            "asOfDate": d(as_date(cfg.get("As of date", date.today().isoformat()))),
            "agingDefinition": str(cfg.get("Aging period definition", "30_60_90_180")),
            "redDotThreshold": red_dot,
            "theme": str(cfg.get("Theme", "blue")).strip().lower(),
            "size": str(cfg.get("Size", "comfortable")).strip().lower(),
        },
        "customers": [{
            "account": c["account"], "name": c["name"], "group": c["group"],
            "terms": c["terms"], "currency": c["currency"], "city": c["city"],
            "address": c["address"], "balance": c["balance"],
            "onTime": c["on_time_pct"], "late": c["late_pct"], "veryLate": c["very_late_pct"],
            "onTimeAmt": c["on_time_amt"], "lateAmt": c["late_amt"], "veryLateAmt": c["very_late_amt"],
            "openInvoices": c["open_invoices"], "lateInvoices": c["late_invoices"],
            "collectionCases": c["collection_cases"], "openActivities": c["open_activities"],
            "lastPaymentAmount": c["last_payment_amount"],
            "lastPaymentDate": d(c["last_payment_date"]),
            "avgInvoiceBalance": c["avg_invoice_balance"],
            "avgAmountLate": c["avg_amount_late"],
            "avgDaysToPay": c["avg_days_to_pay"],
            "creditLimit": c["credit_limit"], "creditUsedPct": c["credit_used_pct"],
            "aged": {k: round(v, 2) for k, v in c["aged"].items()},
            "factors": c["factors"],
        } for c in cust_rows],
        "transactions": [{
            "account": t["account"], "name": t["name"], "onTime": t["on_time"],
            "late": t["late"], "veryLate": t["very_late"], "voucher": t["voucher"],
            "invoice": t["invoice"], "type": t["ttype"], "description": t["description"],
            "date": d(t["date"]), "due": d(t["due"]), "dpd": t["dpd"],
            "amount": t["amount"], "balance": t["balance"], "currency": t["currency"],
        } for t in transactions],
    }


def build_html(payload, out: Path):
    if not TEMPLATE.exists():
        raise SystemExit(f"Template not found: {TEMPLATE}")
    html = TEMPLATE.read_text(encoding="utf-8")
    if "__DEMO_DATA__" not in html:
        raise SystemExit("Template is missing the __DEMO_DATA__ placeholder.")
    blob = json.dumps(payload, separators=(",", ":"), ensure_ascii=False)
    blob = blob.replace("</", "<\\/")            # keep it safe inside <script>
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html.replace("__DEMO_DATA__", blob), encoding="utf-8")


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Build the D365 payment prediction demo.")
    ap.add_argument("--seed", default="airport", choices=sorted(SEED_PACKS))
    ap.add_argument("--force-seed", action="store_true")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument("--out", type=Path, default=DEFAULT_HTML)
    ap.add_argument("--no-html", action="store_true")
    args = ap.parse_args()

    wb_path = args.workbook
    if args.force_seed or not wb_path.exists():
        action = "Reseeded" if wb_path.exists() else "Created"
        seed_workbook(wb_path, args.seed)
        print(f"{action} workbook from '{args.seed}' seed pack -> {wb_path}")

    wb, cfg, customers = read_inputs(wb_path)
    if not customers:
        raise SystemExit("Input_Customers is empty - nothing to generate.")

    red_dot = int(cfg.get("Red dot threshold (%)", 60))
    cust_rows, transactions = generate(customers, cfg)
    write_outputs(wb, wb_path, cfg, cust_rows, transactions, red_dot)

    at_risk = sum(1 for t in transactions if t["on_time"] < red_dot)
    total = sum(t["balance"] for t in transactions)
    print(f"Workbook : {wb_path}")
    print(f"           {len(cust_rows)} customers, {len(transactions)} open transactions, "
          f"{total:,.2f} {cfg.get('Currency','USD')} open")
    print(f"           {at_risk} transactions below the {red_dot}% on-time threshold")

    if not args.no_html:
        build_html(build_payload(cfg, cust_rows, transactions, red_dot), args.out)
        print(f"Dashboard: {args.out}")


if __name__ == "__main__":
    main()
